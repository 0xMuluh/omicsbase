"""Universal file inspector — reads and summarizes uploaded data files."""

from __future__ import annotations

import csv
import io
import json
import logging
import zipfile
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


def inspect_file(file_path: str) -> dict[str, Any]:
    """Inspect a file and return a structured summary.

    Handles: CSV, TSV, Excel, BIOM, QIIME2 .qza, MetaPhlAn profiles, JSON.
    Returns a dict with: format, dimensions, columns, dtypes, preview_rows,
    and any format-specific metadata.
    """
    path = Path(file_path)
    suffix = path.suffix.lower()
    name = path.name.lower()

    try:
        if suffix == ".qza":
            return _inspect_qza(path)
        elif suffix == ".biom":
            return _inspect_biom(path)
        elif suffix in (".csv", ".tsv", ".txt"):
            return _inspect_tabular(path, suffix)
        elif suffix in (".xlsx", ".xls"):
            return _inspect_excel(path)
        elif suffix == ".rds":
            return _inspect_rds(path)
        elif suffix == ".sav":
            return _inspect_sav(path)
        elif suffix == ".json":
            return _inspect_json(path)
        else:
            return {"format": "unknown", "name": path.name, "size_bytes": path.stat().st_size}
    except Exception as e:
        logger.warning("Failed to inspect %s: %s", path, e)
        return {"format": "error", "name": path.name, "error": str(e)}


def _inspect_tabular(path: Path, suffix: str) -> dict[str, Any]:
    """Inspect CSV/TSV/TXT files."""
    sep = "\t" if suffix in (".tsv", ".txt") else ","

    # Sniff the delimiter if .txt
    with open(path, "r") as f:
        sample = f.read(4096)
        if suffix == ".txt":
            sniffer = csv.Sniffer()
            try:
                dialect = sniffer.sniff(sample)
                sep = dialect.delimiter
            except csv.Error:
                sep = "\t"

    # Read with csv
    with open(path, "r") as f:
        reader = csv.reader(f, delimiter=sep)
        rows = []
        for i, row in enumerate(reader):
            rows.append(row)
            if i >= 20:  # header + 20 data rows
                break

    if not rows:
        return {"format": "tabular_empty", "name": path.name}

    header = rows[0]
    data_rows = rows[1:]

    # Count total rows
    with open(path, "r") as f:
        total_rows = sum(1 for _ in f) - 1  # subtract header

    # Detect column types from sample
    col_types = {}
    for col_idx, col_name in enumerate(header):
        values = [r[col_idx] for r in data_rows if col_idx < len(r)]
        col_types[col_name] = _infer_type(values)

    # Detect unique values for low-cardinality columns
    categorical_summary = {}
    if total_rows <= 10000:
        with open(path, "r") as f:
            reader = csv.DictReader(f, delimiter=sep)
            all_values: dict[str, set] = {h: set() for h in header}
            for row in reader:
                for h in header:
                    if h in row and row[h]:
                        all_values[h].add(row[h])

            for col_name, vals in all_values.items():
                if len(vals) <= 20:
                    categorical_summary[col_name] = sorted(vals)

    return {
        "format": "csv" if sep == "," else "tsv",
        "name": path.name,
        "dimensions": {"rows": total_rows, "columns": len(header)},
        "columns": header,
        "column_types": col_types,
        "preview_rows": data_rows[:5],
        "categorical_summary": categorical_summary,
    }


def _inspect_qza(path: Path) -> dict[str, Any]:
    """Inspect QIIME2 .qza artifact — extract metadata from the ZIP structure."""
    result: dict[str, Any] = {"format": "qiime2_qza", "name": path.name}

    try:
        with zipfile.ZipFile(path, "r") as zf:
            names = zf.namelist()

            # Find metadata.yaml inside the archive
            metadata_files = [n for n in names if n.endswith("metadata.yaml")]
            if metadata_files:
                content = zf.read(metadata_files[0]).decode("utf-8")
                result["qiime2_metadata"] = content[:2000]

            # Try to find and read the data file
            data_files = [n for n in names if n.endswith((".biom", ".tsv", ".txt", ".csv"))]
            result["internal_files"] = data_files[:10]

            # If there's a BIOM file inside, note it
            biom_files = [n for n in names if n.endswith(".biom")]
            if biom_files:
                result["contains_biom"] = True
                result["biom_file"] = biom_files[0]

            # If there's a TSV/CSV, read a preview
            tabular_files = [n for n in names if n.endswith((".tsv", ".txt", ".csv"))]
            if tabular_files:
                with zf.open(tabular_files[0]) as inner:
                    content = inner.read().decode("utf-8", errors="replace")
                    lines = content.split("\n")[:20]
                    result["preview_lines"] = lines

    except zipfile.BadZipFile:
        result["error"] = "Not a valid ZIP/QZA file"

    result["size_bytes"] = path.stat().st_size
    return result


def _inspect_biom(path: Path) -> dict[str, Any]:
    """Inspect BIOM format files."""
    result: dict[str, Any] = {"format": "biom", "name": path.name, "size_bytes": path.stat().st_size}

    try:
        import biom

        table = biom.load_table(str(path))
        result["dimensions"] = {
            "observations": table.shape[0],
            "samples": table.shape[1],
        }
        result["sample_ids"] = list(table.ids("sample"))[:10]
        result["observation_ids"] = list(table.ids("observation"))[:10]

        if table.metadata():
            result["has_sample_metadata"] = True
        if table.metadata(axis="observation"):
            result["has_observation_metadata"] = True

    except ImportError:
        result["note"] = "biom-format package not installed; basic inspection only"
    except Exception as e:
        result["error"] = str(e)

    return result


def _inspect_excel(path: Path) -> dict[str, Any]:
    """Inspect Excel files."""
    result: dict[str, Any] = {"format": "excel", "name": path.name, "size_bytes": path.stat().st_size}

    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        result["sheets"] = wb.sheetnames
        preferred_sheet = "SPSS" if "SPSS" in wb.sheetnames else wb.sheetnames[0]

        for sheet_name in wb.sheetnames[:3]:
            ws = wb[sheet_name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                rows.append([str(c) if c is not None else "" for c in row])
                if i >= 10:
                    break

            if rows:
                result[f"sheet_{sheet_name}"] = {
                    "dimensions": {"rows": ws.max_row, "columns": ws.max_column},
                    "preview_rows": rows[:6],
                }

        preferred = wb[preferred_sheet]
        sampled_rows = []
        for index, row in enumerate(preferred.iter_rows(values_only=True)):
            sampled_rows.append(list(row))
            if index >= 200:
                break
        if sampled_rows:
            columns = [
                str(value) if value is not None and str(value).strip() else f"column_{index + 1}"
                for index, value in enumerate(sampled_rows[0])
            ]
            data_rows = sampled_rows[1:]
            result["selected_sheet"] = preferred_sheet
            result["dimensions"] = {"rows": max(preferred.max_row - 1, 0), "columns": preferred.max_column}
            result["columns"] = columns
            result["preview_rows"] = [
                [str(value) if value is not None else "" for value in row]
                for row in data_rows[:5]
            ]
            result["column_types"] = {
                column: _infer_type([
                    "" if index >= len(row) or row[index] is None else str(row[index])
                    for row in data_rows
                ])
                for index, column in enumerate(columns)
            }
            categorical_summary = {}
            for index, column in enumerate(columns):
                values = {
                    str(row[index])
                    for row in data_rows
                    if index < len(row) and row[index] is not None and str(row[index]).strip()
                }
                if 1 < len(values) <= 20:
                    categorical_summary[column] = sorted(values)
            result["categorical_summary"] = categorical_summary

        wb.close()

    except Exception as e:
        result["error"] = str(e)

    return result


def _inspect_rds(path: Path) -> dict[str, Any]:
    """Basic inspection of R .rds files — just size and type detection."""
    return {
        "format": "rds",
        "name": path.name,
        "size_bytes": path.stat().st_size,
        "note": "RDS files are read natively in R; the generated data.R will use readRDS().",
    }


def _inspect_sav(path: Path, *, max_preview_rows: int = 5) -> dict[str, Any]:
    """Inspect SPSS .sav files via R haven when available."""
    result: dict[str, Any] = {
        "format": "spss",
        "name": path.name,
        "size_bytes": path.stat().st_size,
        "editable": False,
    }
    try:
        import json
        import subprocess

        r_path = str(path).replace("\\", "\\\\").replace('"', '\\"')
        script = f"""
if (!requireNamespace("haven", quietly = TRUE) || !requireNamespace("jsonlite", quietly = TRUE)) {{
  stop("haven and jsonlite are required to preview SPSS files")
}}
df <- as.data.frame(haven::read_sav("{r_path}"), check.names = FALSE)
n_preview <- min({int(max_preview_rows)}, nrow(df))
char_df <- as.data.frame(lapply(df, function(col) {{
  values <- as.character(col)
  values[is.na(col)] <- ""
  values
}}), stringsAsFactors = FALSE, check.names = FALSE)
types <- vapply(df, function(col) {{
  if (inherits(col, "haven_labelled")) "labelled"
  else if (is.numeric(col)) "numeric"
  else if (is.logical(col)) "logical"
  else "text"
}}, character(1))
cat(jsonlite::toJSON(list(
  dimensions = list(rows = nrow(df), columns = ncol(df)),
  columns = names(df),
  column_types = as.list(types),
  preview_rows = head(char_df, n_preview),
  preview_truncated = nrow(df) > n_preview
), dataframe = "values", auto_unbox = TRUE, null = "null"))
"""
        from app.services.runner import run_command_sync

        success, run_output = run_command_sync(
            ["Rscript", "-e", script],
            cwd=str(path.parent),
            timeout=60,
        )
        if not success or not run_output.strip():
            result["note"] = "SPSS preview unavailable; file will be read with haven during analysis."
            result["error"] = run_output.strip()[:500]
            return result
        payload = json.loads(run_output.strip().splitlines()[-1])
        result.update(payload)
        preview = result.get("preview_rows") or []
        columns = result.get("columns") or []
        if preview and columns and isinstance(preview[0], dict):
            result["preview_rows"] = [
                [str(row.get(col, "")) for col in columns]
                for row in preview
            ]
    except FileNotFoundError:
        result["note"] = "Rscript not available; SPSS files are read with haven during analysis."
    except Exception as exc:
        result["note"] = "SPSS preview failed; file will be read with haven during analysis."
        result["error"] = str(exc)
    return result


TABULAR_PREVIEW_EXTENSIONS = {".csv", ".tsv", ".txt", ".xlsx", ".xls", ".sav"}


def preview_tabular_file(file_path: str, *, max_rows: int = 100) -> dict[str, Any]:
    """Return a UI-oriented tabular preview for CSV/TSV/Excel/SPSS files."""
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix not in TABULAR_PREVIEW_EXTENSIONS:
        return {
            "format": "unsupported",
            "name": path.name,
            "error": f"No tabular preview for {suffix or 'unknown'} files",
        }

    if suffix in {".csv", ".tsv", ".txt"}:
        summary = _inspect_tabular_preview(path, suffix, max_rows=max_rows)
        summary["editable"] = True
        return summary
    if suffix in {".xlsx", ".xls"}:
        summary = _inspect_excel_preview(path, max_rows=max_rows)
        summary["editable"] = False
        return summary
    summary = _inspect_sav(path, max_preview_rows=max_rows)
    summary["editable"] = False
    return summary


def _inspect_tabular_preview(path: Path, suffix: str, *, max_rows: int) -> dict[str, Any]:
    """CSV/TSV preview with a larger row window for the workspace table viewer."""
    summary = _inspect_tabular(path, suffix)
    sep = "\t" if summary.get("format") == "tsv" else ","
    if suffix == ".txt":
        with open(path, "r") as f:
            sample = f.read(4096)
        try:
            sep = csv.Sniffer().sniff(sample).delimiter
        except csv.Error:
            sep = "\t"

    with open(path, "r") as f:
        reader = csv.reader(f, delimiter=sep)
        rows = []
        for i, row in enumerate(reader):
            rows.append(row)
            if i >= max_rows:  # header + max_rows data rows
                break

    if not rows:
        return summary

    header = rows[0]
    data_rows = rows[1:]
    summary["columns"] = header
    summary["preview_rows"] = data_rows
    summary["preview_truncated"] = summary.get("dimensions", {}).get("rows", 0) > len(data_rows)
    return summary


def _inspect_excel_preview(path: Path, *, max_rows: int) -> dict[str, Any]:
    """Excel preview with a larger row window for the workspace table viewer."""
    summary = _inspect_excel(path)
    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheet_name = summary.get("selected_sheet") or (wb.sheetnames[0] if wb.sheetnames else None)
        if not sheet_name:
            wb.close()
            return summary
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append([str(c) if c is not None else "" for c in row])
            if i >= max_rows:
                break
        wb.close()
        if rows:
            header = [
                cell if cell.strip() else f"column_{idx + 1}"
                for idx, cell in enumerate(rows[0])
            ]
            data_rows = rows[1:]
            summary["columns"] = header
            summary["preview_rows"] = data_rows
            dims = summary.get("dimensions") or {}
            summary["preview_truncated"] = int(dims.get("rows") or 0) > len(data_rows)
    except Exception as exc:
        summary["error"] = str(exc)
    return summary


def _inspect_json(path: Path) -> dict[str, Any]:
    """Inspect JSON files."""
    result: dict[str, Any] = {"format": "json", "name": path.name, "size_bytes": path.stat().st_size}
    try:
        with open(path) as f:
            data = json.load(f)
        if isinstance(data, dict):
            result["top_level_keys"] = list(data.keys())[:20]
        elif isinstance(data, list):
            result["array_length"] = len(data)
            if data:
                result["first_item_keys"] = list(data[0].keys())[:20] if isinstance(data[0], dict) else None
    except Exception as e:
        result["error"] = str(e)
    return result


def _infer_type(values: list[str]) -> str:
    """Infer column type from a sample of string values."""
    if not values:
        return "empty"

    numeric_count = 0
    int_count = 0
    for v in values:
        if not v:
            continue
        try:
            f = float(v)
            numeric_count += 1
            if f == int(f):
                int_count += 1
        except (ValueError, OverflowError):
            pass

    non_empty = [v for v in values if v]
    if not non_empty:
        return "empty"

    if numeric_count == len(non_empty):
        return "integer" if int_count == numeric_count else "numeric"

    unique = set(non_empty)
    if len(unique) <= 10:
        return "categorical"

    return "text"


def format_file_summary_for_llm(summary: dict) -> str:
    """Format a file inspection summary as readable text for the LLM."""
    parts = [f"File: {summary.get('name', 'unknown')}"]
    parts.append(f"Format: {summary.get('format', 'unknown')}")

    if "dimensions" in summary:
        d = summary["dimensions"]
        if "rows" in d:
            parts.append(f"Dimensions: {d['rows']} rows × {d['columns']} columns")
        elif "observations" in d:
            parts.append(f"Dimensions: {d['observations']} observations × {d['samples']} samples")

    if "columns" in summary:
        parts.append(f"Columns: {', '.join(summary['columns'])}")

    if "column_types" in summary:
        type_info = [f"  {col}: {dtype}" for col, dtype in summary["column_types"].items()]
        parts.append("Column types:\n" + "\n".join(type_info))

    if "categorical_summary" in summary:
        for col, vals in summary["categorical_summary"].items():
            # Emit cardinality count instead of raw string values to preserve privacy
            parts.append(f"  {col}: {len(vals)} unique category levels")

    if "preview_rows" in summary:
        # SECURITY: Do NOT send raw cell values to LLM — emit schema note only.
        row_count = len(summary["preview_rows"])
        parts.append(f"Preview: {row_count} sample row(s) available (schema above, raw values withheld from LLM)")

    if "preview_lines" in summary:
        # Omit raw text lines for data files; indicate preview availability
        line_count = len(summary["preview_lines"])
        parts.append(f"Preview lines: {line_count} initial line(s) inspected")

    if "qiime2_metadata" in summary:
        parts.append("QIIME2 metadata format detected")

    return "\n".join(parts)
