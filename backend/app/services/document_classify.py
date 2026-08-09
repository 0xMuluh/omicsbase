"""Agentic, privacy-aware classification of uploaded study documents.

The agent inspects every upload to decide what it is: a plan document, a
feature table, metadata, or supporting material. Privacy rule: typical data
files expose only headers/columns/dimensions to the LLM; plan documents are
the only files whose full text is consumed. Everything is size-capped.
"""

from __future__ import annotations

import json
import logging
import re
import zipfile
from pathlib import Path

from app.services.llm import call_llm, resolve_target

logger = logging.getLogger(__name__)

# Size caps (configurable via settings later if needed).
MAX_INSPECT_BYTES = 10 * 1024 * 1024      # files larger than this are not content-read
MAX_DOC_TEXT_CHARS = 50_000               # extracted plan text cap
MAX_CLASSIFIER_PREVIEW_CHARS = 6_000      # doc preview sent to the classifier
MAX_AMBIGUOUS_PREVIEW_CHARS = 2_000       # txt preview sent to the classifier

# Typical data extensions: only headers are ever shown for these.
DATA_EXTENSIONS = {
    ".csv", ".tsv", ".txt", ".xlsx", ".xls", ".sav", ".rds", ".biom",
    ".qza", ".qza2", ".parquet", ".json", ".feather",
}
# Document-like extensions: full extracted text may be inspected.
DOC_EXTENSIONS = {
    ".docx", ".doc", ".pdf", ".md", ".markdown", ".pptx", ".html", ".htm", ".rtf", ".odt",
}

ROLE_PLAN = "analysis_plan"
ROLE_FEATURE = "feature_table"
ROLE_METADATA = "metadata"
ROLE_SUPPORT = "supporting"

VALID_ROLES = {ROLE_PLAN, ROLE_FEATURE, ROLE_METADATA, ROLE_SUPPORT}


def extract_document_text(path: str, max_chars: int = MAX_DOC_TEXT_CHARS) -> str:
    """Extract readable text from any document, size-capped, with graceful fallbacks."""
    try:
        size = Path(path).stat().st_size
    except OSError:
        return ""
    if size <= 0 or size > MAX_INSPECT_BYTES:
        return ""

    suffix = Path(path).suffix.lower()
    text = ""
    if suffix == ".docx":
        text = _extract_docx(path)
    elif suffix == ".pdf":
        text = _extract_pdf(path)
    elif suffix in {".txt", ".md", ".markdown", ".rtf", ".html", ".htm"}:
        text = _read_text(path)
    elif suffix == ".xlsx":
        text = _extract_xlsx_headers(path)
    elif suffix in {".csv", ".tsv"}:
        text = _extract_delimited_headers(path, "," if suffix == ".csv" else "\t")
    elif suffix == ".odt":
        text = _extract_odt(path)
    if not text:
        return ""
    return text[:max_chars]


def _read_text(path: str, max_chars: int = MAX_DOC_TEXT_CHARS) -> str:
    for encoding in ("utf-8", "latin-1"):
        try:
            return Path(path).read_text(encoding=encoding, errors="replace")[:max_chars]
        except (UnicodeDecodeError, OSError):
            continue
    return ""


def _extract_docx(path: str) -> str:
    """Extract docx text via python-docx if installed, else stdlib zip+xml."""
    try:
        import docx  # type: ignore

        document = docx.Document(path)
        return "\n".join(p.text for p in document.paragraphs if p.text)
    except Exception:
        pass
    try:
        with zipfile.ZipFile(path) as archive:
            xml = archive.read("word/document.xml").decode("utf-8", errors="replace")
        xml = re.sub(r"<w:tab[^>]*/>", "\t", xml)
        xml = re.sub(r"</w:p>", "\n", xml)
        xml = re.sub(r"<[^>]+>", "", xml)
        return re.sub(r"\n{3,}", "\n\n", xml).strip()
    except Exception as exc:
        logger.debug("docx extraction failed for %s: %s", path, exc)
        return ""


def _extract_pdf(path: str) -> str:
    """Extract PDF text via pypdf if installed, else the pdftotext binary."""
    try:
        import pypdf  # type: ignore

        reader = pypdf.PdfReader(path)
        return "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception:
        pass
    import shutil

    if shutil.which("pdftotext"):
        import subprocess

        try:
            result = subprocess.run(
                ["pdftotext", path, "-"],
                capture_output=True,
                text=True,
                timeout=60,
                check=False,
            )
            return result.stdout
        except Exception as exc:
            logger.debug("pdftotext failed for %s: %s", path, exc)
    return ""


def _extract_xlsx_headers(path: str) -> str:
    """Header-only summary for xlsx — never full cell contents (privacy)."""
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        parts = [f"Sheets: {', '.join(workbook.sheetnames)}"]
        for sheet_name in workbook.sheetnames[:3]:
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(min_row=1, max_row=3, values_only=True)
            headers = None
            for row in rows:
                if row and any(v is not None for v in row):
                    headers = [str(v) for v in row if v is not None]
                    break
            if headers:
                parts.append(f"{sheet_name} headers: {', '.join(headers)}")
        return "\n".join(parts)
    except Exception as exc:
        logger.debug("xlsx header extraction failed for %s: %s", path, exc)
        return ""


def _extract_delimited_headers(path: str, delimiter: str) -> str:
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as handle:
            header_line = handle.readline().strip()
        if not header_line:
            return ""
        return "Header: " + header_line[:2000]
    except OSError:
        return ""


def _extract_odt(path: str) -> str:
    try:
        with zipfile.ZipFile(path) as archive:
            xml = archive.read("content.xml").decode("utf-8", errors="replace")
        xml = re.sub(r"</text:p>", "\n", xml)
        xml = re.sub(r"<[^>]+>", "", xml)
        return re.sub(r"\n{3,}", "\n\n", xml).strip()
    except Exception:
        return ""


def _classifier_input(files: list[dict]) -> list[dict]:
    """Build the privacy-aware classifier payload for each file."""
    payload = []
    for item in files:
        name = str(item.get("name") or "unknown")
        suffix = Path(name).suffix.lower()
        extracted = str(item.get("text") or "")
        if suffix in DATA_EXTENSIONS and suffix not in {".txt", ".csv", ".tsv"}:
            preview = extracted[:MAX_CLASSIFIER_PREVIEW_CHARS]  # headers only by construction
            show_content = False
        elif suffix in DOC_EXTENSIONS:
            preview = extracted[:MAX_CLASSIFIER_PREVIEW_CHARS]
            show_content = True
        else:
            preview = extracted[:MAX_AMBIGUOUS_PREVIEW_CHARS]
            show_content = True
        payload.append({
            "file": name,
            "size_bytes": int(item.get("size_bytes") or 0),
            "extension": suffix,
            "content_preview": preview if show_content else f"(data file — headers only) {preview[:800]}",
            "full_content_inspectable": show_content,
        })
    return payload


CLASSIFIER_SYSTEM = """You classify uploaded files for a scientific omics analysis project.

Return JSON: {"files": [{"file": "<name>", "role": "analysis_plan"|"feature_table"|"metadata"|"supporting", "reason": "<one line>", "is_plan": true|false}]}

Rules:
- A plan document (draft analysis plan, protocol, study brief, pasted plan) is "analysis_plan" with is_plan=true. Plans are OK to be fully read.
- Feature tables are "feature_table"; sample/clinical tables are "metadata"; other support material is "supporting".
- Data files expose only headers; never guess that data rows are a plan.
- A .txt with plan-like prose is a plan; a .txt that looks tabular (reports, counts, taxonomy) is data.
- Reason from the file name, extension, and preview; never invent content."""


async def classify_uploads(files: list[dict]) -> dict[str, dict]:
    """Classify uploaded files with one LLM call.

    ``files``: list of {"name", "size_bytes", "text" (extracted, capped)}.
    Returns {name: {"role": str, "is_plan": bool, "reason": str}}.
    No heuristic fallback: when the LLM is unavailable, no roles are assigned
    and existing roles stay unchanged.
    """
    payload = _classifier_input(files)
    provider, model = resolve_target("fast")
    try:
        response = await call_llm(
            system_prompt=CLASSIFIER_SYSTEM,
            user_prompt=json.dumps(payload),
            response_format="json",
            max_tokens=1500,
            model_override=model,
            provider_override=provider,
        )
        data = json.loads(response)
        results: dict[str, dict] = {}
        for entry in data.get("files", []):
            name = entry.get("file")
            role = entry.get("role")
            if name and role in VALID_ROLES:
                results[name] = {
                    "role": role,
                    "is_plan": bool(entry.get("is_plan")) or role == ROLE_PLAN,
                    "reason": str(entry.get("reason") or "llm"),
                }
        return results
    except Exception as exc:
        logger.warning("Upload classification failed; roles left unchanged: %s", exc)
        return {}


async def agentic_plan_sources(files: list) -> list[str]:
    """Classify uploads and return extracted plan text for the planner.

    ``files``: UploadedFile ORM records. Best-effort: classification failures
    never block planning. Plan text is size-capped.
    """
    plans: list[str] = []
    if not files:
        return plans

    extracted = []
    for file_record in files:
        path = getattr(file_record, "file_path", None)
        name = getattr(file_record, "name", None) or (Path(path).name if path else "unknown")
        text = extract_document_text(str(path)) if path else ""
        extracted.append({"name": name, "size_bytes": Path(path).stat().st_size if path and Path(path).exists() else 0, "text": text})

    classified = await classify_uploads(extracted)

    for file_record, meta in zip(files, extracted):
        name = meta["name"]
        decision = classified.get(name)
        if not decision:
            continue
        role = decision["role"]
        if role == ROLE_PLAN or decision.get("is_plan"):
            if role != getattr(file_record, "file_role", None):
                logger.info("Reclassifying %s: %s -> %s", name, file_record.file_role, role)
                file_record.file_role = role
            if meta["text"]:
                plans.append(f"### Plan document: {name}\n\n{meta['text']}")
        elif role != getattr(file_record, "file_role", None):
            logger.info("Reclassifying %s: %s -> %s", name, file_record.file_role, role)
            file_record.file_role = role
    return plans
