"""Materialize deterministic analysis recipes into generated projects."""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

import yaml

from app.schemas.schemas import AnalysisPlan
from app.services.recipe_registry import get_recipe, resolve_recipe


TEMPLATE_ROOT = Path(__file__).resolve().parents[1] / "recipe_templates"
SUPPORTED_TABULAR_EXTENSIONS = {".csv", ".tsv", ".txt", ".xlsx", ".xls", ".sav", ".rds"}
RECIPE_TEMPLATE_BINDINGS = {
    "microbiome.alpha_diversity": ("microbiome/alpha.qmd", "code/alpha/alpha.qmd"),
    "microbiome.beta_diversity": ("microbiome/beta.qmd", "code/beta/beta.qmd"),
    "microbiome.permanova": ("microbiome/permanova.qmd", "code/beta/permanova.qmd"),
    "microbiome.limrots_differential_abundance": (
        "microbiome/limrots.qmd",
        "code/daa/daa_limrots.qmd",
    ),
    "metabolomics.linear_feature_scan": (
        "metabolomics/limma_panel.qmd",
        "code/primary/metabolite_panel.qmd",
    ),
    "metabolomics.repeated_measures_mixed_model": (
        "metabolomics/longitudinal.qmd",
        "code/primary/longitudinal_models.qmd",
    ),
}


def materialize_recipe_project(
    project_dir: str,
    plan: AnalysisPlan,
    study_manifest: dict[str, Any] | None,
    uploaded_file_paths: dict[str, list[str]],
) -> dict[str, Any]:
    """Write deterministic recipe files and return their workflow bindings."""
    manifest = study_manifest or {}
    domain = plan.domain if plan.domain in {"microbiome", "metabolomics"} else manifest.get("domain")
    if domain not in {"microbiome", "metabolomics"}:
        return {"domain": "unknown", "files": [], "step_paths": {}, "recipe_ids": []}

    feature_source = _select_feature_source(domain, uploaded_file_paths)
    if not feature_source or Path(feature_source).suffix.lower() not in SUPPORTED_TABULAR_EXTENSIONS:
        return {"domain": domain, "files": [], "step_paths": {}, "recipe_ids": []}

    metadata_source = _first_path(uploaded_file_paths.get("metadata"))
    if domain == "metabolomics" and not metadata_source:
        metadata_source = feature_source

    selected_steps: list[tuple[Any, dict[str, Any]]] = []
    for step in plan.workflow:
        if not step.enabled:
            continue
        recipe = get_recipe(step.recipe_id) if step.recipe_id else None
        recipe = recipe or resolve_recipe(step.id, domain)
        if recipe:
            selected_steps.append((step, recipe))

    inventory_recipe_id = f"{domain}.inventory"
    requested_recipe_ids = [inventory_recipe_id]
    requested_recipe_ids.extend(
        recipe["id"] for _, recipe in selected_steps if recipe["id"] not in requested_recipe_ids
    )
    recipe_ids = _expand_recipe_dependencies(requested_recipe_ids)

    config = _build_study_config(
        plan=plan,
        manifest=manifest,
        domain=domain,
        feature_source=feature_source,
        metadata_source=metadata_source,
        recipe_ids=recipe_ids,
    )

    base = Path(project_dir)
    code_dir = base / "code"
    code_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    step_paths: dict[str, str] = {}

    written.append(_copy_template(TEMPLATE_ROOT / "common" / "recipe_runtime.R", code_dir / "recipe_runtime.R"))
    written.append(_write_text(code_dir / "funct.R", 'source("recipe_runtime.R")\n'))
    written.append(_write_text(code_dir / "study_config.yml", yaml.safe_dump(config, sort_keys=False)))
    written.append(_copy_template(TEMPLATE_ROOT / domain / "data.R", code_dir / "data.R"))
    inventory_path = _copy_template(
        TEMPLATE_ROOT / "common" / "inventory.qmd",
        code_dir / "data" / "data_summary.qmd",
    )
    written.append(inventory_path)

    recipe_page_paths = {
        inventory_recipe_id: "code/data/data_summary.qmd",
    }
    for recipe_id, (_, project_path) in RECIPE_TEMPLATE_BINDINGS.items():
        if recipe_id in recipe_ids:
            continue
        stale_path = base / project_path
        if stale_path.exists():
            stale_path.unlink()
    for recipe_id in recipe_ids:
        binding = RECIPE_TEMPLATE_BINDINGS.get(recipe_id)
        if not binding:
            continue
        template_path, project_path = binding
        written.append(_copy_template(TEMPLATE_ROOT / template_path, base / project_path))
        recipe_page_paths[recipe_id] = project_path

    for step, recipe in selected_steps:
        project_path = recipe_page_paths.get(recipe["id"])
        if project_path:
            step_paths[step.id] = project_path

    return {
        "domain": domain,
        "files": [str(path) for path in written],
        "step_paths": step_paths,
        "recipe_ids": recipe_ids,
        "config": config,
    }


def recipe_page_path(recipe_id: str) -> str | None:
    if recipe_id.endswith(".inventory"):
        return "data/data_summary.qmd"
    binding = RECIPE_TEMPLATE_BINDINGS.get(recipe_id)
    return binding[1].removeprefix("code/") if binding else None


def _build_study_config(
    *,
    plan: AnalysisPlan,
    manifest: dict[str, Any],
    domain: str,
    feature_source: str,
    metadata_source: str | None,
    recipe_ids: list[str],
) -> dict[str, Any]:
    feature_record = _manifest_file(manifest, feature_source)
    metadata_record = _manifest_file(manifest, metadata_source)
    feature_id = _select_feature_identifier(feature_record)
    sample_id = _select_identifier(
        manifest,
        metadata_record,
        feature_source=feature_source,
        metadata_source=metadata_source,
        feature_id=feature_id,
    )
    visit = _select_visit_column(metadata_record)
    recipe_parameters = _collect_recipe_parameters(plan, recipe_ids)
    analysis_parameters = _collect_analysis_parameters(recipe_parameters)

    return {
        "schema_version": "1.0",
        "study": {
            "title": plan.project_name,
            "domain": domain,
            "question": plan.question,
        },
        "paths": {
            "feature_table": f"../data/{Path(feature_source).name}",
            "metadata": f"../data/{Path(metadata_source).name}" if metadata_source else None,
        },
        "identifiers": {
            "sample_id": sample_id,
            "subject_id": sample_id,
            "feature_id": feature_id,
            "visit": visit,
        },
        "features": {
            "orientation": "auto" if domain == "microbiome" else "samples_by_features",
            "input_scale": "auto",
            "columns": [],
        },
        "variables": {
            "grouping": plan.grouping_variable,
            "group_levels": plan.group_levels,
            "covariates": plan.covariates,
        },
        "analyses": {
            "recipes": recipe_ids,
            "fdr_method": "BH",
            "random_seed": 20260730,
            "permutations": 999,
            "clr_pseudocount": 0.000001,
            "limrots_bootstrap_iterations": 200,
            "min_complete_cases": 20,
            "lmm_min_complete_cases": 30,
            **analysis_parameters,
            "recipe_parameters": recipe_parameters,
        },
        "provenance": {
            "manifest_version": manifest.get("version"),
            "recipe_registry_version": plan.recipe_registry_version,
        },
    }


def _collect_recipe_parameters(
    plan: AnalysisPlan,
    recipe_ids: list[str],
) -> dict[str, dict[str, Any]]:
    parameters: dict[str, dict[str, Any]] = {}
    steps_by_recipe = {
        step.recipe_id: step
        for step in plan.workflow
        if step.enabled and step.recipe_id
    }
    for recipe_id in recipe_ids:
        recipe = get_recipe(recipe_id)
        if not recipe:
            continue
        recipe_values = dict(recipe.get("parameters") or {})
        step = steps_by_recipe.get(recipe_id)
        if step and step.parameters:
            recipe_values.update(step.parameters)
        parameters[recipe_id] = recipe_values
    return parameters


def _collect_analysis_parameters(
    recipe_parameters: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    parameters: dict[str, Any] = {}
    for values in recipe_parameters.values():
        parameters.update(values)

    aliases = {
        "seed": "random_seed",
        "bootstrap_iterations": "limrots_bootstrap_iterations",
        "pseudocount": "clr_pseudocount",
    }
    for source, target in aliases.items():
        if source in parameters:
            parameters[target] = parameters[source]
    return parameters


def _select_feature_source(domain: str, paths: dict[str, list[str]]) -> str | None:
    explicit = _first_path(paths.get("feature_table"))
    if explicit:
        return explicit
    if domain == "metabolomics":
        other = _first_path(paths.get("other"))
        if other:
            return other
    for role, role_paths in paths.items():
        if role not in {"analysis_plan", "metadata", "taxonomy"} and role_paths:
            return role_paths[0]
    return None


def _select_identifier(
    manifest: dict[str, Any],
    metadata_record: dict[str, Any] | None,
    *,
    feature_source: str | None = None,
    metadata_source: str | None = None,
    feature_id: str | None = None,
) -> str | None:
    """Pick the sample identifier column.

    Name heuristics score candidates; when the feature table and metadata
    file are readable, candidates whose actual values overlap the feature
    table's sample columns get a decisive boost (this catches layouts where
    ``id`` is a re-encoded key and ``sample`` holds the matching IDs).
    """
    metadata_name = (metadata_record or {}).get("name")
    candidates = manifest.get("identifier_candidates", [])
    preferred = []
    for candidate in candidates:
        column = str(candidate.get("column", "") or "")
        if not column:
            continue
        score = 0
        if candidate.get("file") == metadata_name:
            score += 3
        normalized = column.lower().replace(" ", "").replace("_", "")
        if normalized in {"sampleid", "studyid", "participantid", "subjectid", "id"}:
            score += 5
        elif normalized.endswith("id"):
            score += 2
        score += _overlap_evidence(metadata_source, column, feature_source, feature_id)
        preferred.append((score, column))
    preferred.sort(key=lambda item: item[0], reverse=True)
    if preferred and preferred[0][0] > 0:
        return preferred[0][1]
    columns = (metadata_record or {}).get("columns") or []
    if metadata_source and feature_source:
        best_column, best_score = None, 0
        for column in columns:
            evidence = _overlap_evidence(metadata_source, str(column), feature_source, feature_id)
            if evidence > best_score:
                best_column, best_score = column, evidence
        if best_column is not None and best_score > 0:
            return best_column
    return columns[0] if columns else None


def _overlap_evidence(
    metadata_source: str | None,
    column: str,
    feature_source: str | None,
    feature_id: str | None,
) -> int:
    """Score how many metadata values match the feature table's sample columns."""
    if not metadata_source or not feature_source or not column:
        return 0
    sample_columns = _read_sample_columns(feature_source, feature_id)
    if not sample_columns:
        return 0
    values = _read_column_values(metadata_source, column, limit=100)
    if not values:
        return 0
    overlap = sum(1 for value in values if value in sample_columns)
    if overlap == 0:
        return 0
    # Decisive when a meaningful fraction of values align; capped otherwise.
    return 10 if overlap >= min(3, len(sample_columns)) else 4


def _read_sample_columns(source: str, feature_id: str | None) -> set[str] | None:
    """Return the sample column names of a feature table header (best effort)."""
    try:
        suffix = Path(source).suffix.lower()
        if suffix in {".csv", ".tsv", ".txt"}:
            delimiter = "," if suffix == ".csv" else "\t"
            with open(source, "r", encoding="utf-8", errors="replace") as handle:
                header = handle.readline().strip()
            columns = [col.strip() for col in header.split(delimiter) if col.strip()]
        elif suffix in {".xlsx", ".xls"}:
            from openpyxl import load_workbook

            workbook = load_workbook(source, read_only=True, data_only=True)
            sheet = workbook[workbook.sheetnames[0]]
            columns = []
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                columns = [str(value) for value in row if value is not None]
        else:
            return None
        if feature_id and feature_id in columns:
            columns.remove(feature_id)
        elif columns and not feature_id:
            columns = columns[1:]
        return {column for column in columns if column and column.lower() not in {"sample_id", "subject_id"}}
    except Exception:
        return None


def _read_column_values(source: str, column: str, limit: int = 100) -> set[str]:
    """Read up to ``limit`` values of one column from a tabular file (best effort)."""
    values: set[str] = set()
    try:
        suffix = Path(source).suffix.lower()
        if suffix in {".csv", ".tsv", ".txt"}:
            delimiter = "," if suffix == ".csv" else "\t"
            with open(source, "r", encoding="utf-8", errors="replace") as handle:
                lines = handle.readlines()
            if not lines:
                return values
            header = [cell.strip() for cell in lines[0].split(delimiter)]
            try:
                index = header.index(column)
            except ValueError:
                return values
            for line in lines[1 : limit + 1]:
                cells = [cell.strip() for cell in line.split(delimiter)]
                if index < len(cells) and cells[index]:
                    values.add(cells[index])
        elif suffix in {".xlsx", ".xls"}:
            from openpyxl import load_workbook

            workbook = load_workbook(source, read_only=True, data_only=True)
            sheet = workbook[workbook.sheetnames[0]]
            index = None
            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index is None:
                    for i, value in enumerate(row):
                        if str(value) == column:
                            index = i
                            break
                    continue
                if index is not None and row_index - 1 >= limit:
                    break
                if index < len(row) and row[index] is not None:
                    values.add(str(row[index]).strip())
    except Exception:
        pass
    return values


def _select_feature_identifier(feature_record: dict[str, Any] | None) -> str | None:
    columns = (feature_record or {}).get("columns") or []
    for column in columns:
        normalized = str(column).lower()
        if any(term in normalized for term in ("feature", "taxon", "otu", "asv", "metabolite")):
            return column
    return columns[0] if columns else None


def _select_visit_column(metadata_record: dict[str, Any] | None) -> str | None:
    for column in (metadata_record or {}).get("columns") or []:
        normalized = str(column).lower()
        if "visit" in normalized or normalized in {"time", "timepoint", "wave"}:
            return column
    return None


def _manifest_file(manifest: dict[str, Any], source: str | None) -> dict[str, Any] | None:
    if not source:
        return None
    name = Path(source).name
    return next((item for item in manifest.get("files", []) if item.get("name") == name), None)


def _first_path(paths: list[str] | None) -> str | None:
    return paths[0] if paths else None


def _expand_recipe_dependencies(recipe_ids: list[str]) -> list[str]:
    expanded: list[str] = []

    def add(recipe_id: str) -> None:
        recipe = get_recipe(recipe_id)
        if not recipe:
            return
        for dependency in recipe.get("depends_on", []):
            add(dependency)
        if recipe_id not in expanded:
            expanded.append(recipe_id)

    for recipe_id in recipe_ids:
        add(recipe_id)
    return expanded


def _copy_template(source: Path, target: Path) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(source, target)
    return target


def _write_text(path: Path, content: str) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content)
    return path
